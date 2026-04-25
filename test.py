import re
import sys
import hashlib
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _normalize_spaces(s: str) -> str:
    s = re.sub(r'[ \t]+', ' ', s)
    s = re.sub(r'\s*\n\s*', ' ', s)
    return s.strip()

def _remove_mentions(s: str) -> str:
    # @⁨Name⁩ ve @Name mention'larını sil
    s = re.sub(r'@⁨[^⁩]+⁩', ' ', s)
    s = re.sub(r'@\+?\d[\d\s().-]{6,}\d', ' ', s)  # @+90...
    s = re.sub(r'@\S+', ' ', s)
    return s

def _remove_phones(s: str) -> str:
    # TR telefon formatlarını azalt
    s = re.sub(r'\+90[\s().-]*5\d{2}[\s().-]*\d{3}[\s().-]*\d{2}[\s().-]*\d{2}', ' ', s)
    s = re.sub(r'\b0?5\d{9}\b', ' ', s)
    return s

def _remove_noise_tokens(s: str) -> str:
    # sd/dslam/mar/arm gibi operasyonel tokenları azalt (lokal önemli olduğu için tut)
    def repl(m):
        return m.group(0) if m.group(0).lower() == "lokal" else " "
    s = re.sub(r'\b(?:sd|dslam|mar|lokal|arm)\b', repl, s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{1,3}\s*sd\b', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\b[a-z]\d{1,3}\b', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{1,3}[a-z]\b', ' ', s, flags=re.IGNORECASE)
    return s

def build_clean_description(raw_text: str, service_numbers: list[str]) -> str:
    s = raw_text or ""
    for n in service_numbers or []:
        s = s.replace(n, " ")
    s = _remove_mentions(s)
    s = _remove_phones(s)
    s = _remove_noise_tokens(s)
    s = _normalize_spaces(s)
    return s

def build_rule_summary(clean_text: str, mentions: list[str] | None = None) -> str:
    mentions = mentions or []
    t = (clean_text or "").lower()

    def any_(*words):
        return any(w in t for w in words)

    def _extract_day_phrase(text: str) -> str:
        """
        Metinden "pazar günü", "yarın", "ertesi gün" gibi gün/zaman ifadesini yakalar.
        Eğer yakınında "istiyor" (istiyro/istiyo gibi yazımlar dahil) geçiyorsa
        birleşik şekilde "Pazar günü istiyor" gibi döndürür.
        Uydurma yok: sadece metinde geçeni, normalize edip döndürür.
        """
        if not text:
            return ""
        s = text
        sl = text.lower()

        def norm_want(v: str) -> str:
            # Yaygın yazım hatalarını normalize et (sadece output için)
            v = re.sub(r'\bistiyro\b', 'istiyor', v, flags=re.IGNORECASE)
            v = re.sub(r'\bistiyo\b', 'istiyor', v, flags=re.IGNORECASE)
            v = re.sub(r'\bistiyr\b', 'istiyor', v, flags=re.IGNORECASE)
            return v

        # Göreli ifadeler
        rel = [
            r"\bbugün\b",
            r"\byarın\b",
            r"\byarin\b",
            r"\böbür gün\b",
            r"\bobur gun\b",
            r"\bert(es)?i gün\b",
            r"\bertesi gun\b",
        ]
        for pat in rel:
            m = re.search(pat, sl, flags=re.IGNORECASE)
            if m:
                v0 = m.group(0).strip()
                v = v0[:1].upper() + v0[1:]
                # Yakın çevrede "istiyor" var mı?
                win = sl[max(0, m.start() - 20): m.end() + 20]
                if re.search(r'\bisti(?:yor|yo|yro|yr)\b', win, flags=re.IGNORECASE):
                    return norm_want(f"{v} istiyor")
                return v

        # Haftanın günleri (opsiyonel "günü/gunu" ile)
        m = re.search(
            r"\b(pazartesi|salı|sali|çarşamba|carsamba|perşembe|persembe|cuma|cumartesi|pazar)\b(?:\s+gün[üu])?\b",
            sl,
            flags=re.IGNORECASE,
        )
        if not m:
            return ""

        day = m.group(1).lower()
        day_map = {
            "pazartesi": "Pazartesi",
            "salı": "Salı",
            "sali": "Salı",
            "çarşamba": "Çarşamba",
            "carsamba": "Çarşamba",
            "perşembe": "Perşembe",
            "persembe": "Perşembe",
            "cuma": "Cuma",
            "cumartesi": "Cumartesi",
            "pazar": "Pazar",
        }
        nice = day_map.get(day, day[:1].upper() + day[1:])
        # Metinde 'günü/gunu' varsa özellikle koru, yoksa kullanıcı diline uygun ekle
        tail = m.group(0).lower()
        base = f"{nice} günü" if "gün" in tail else nice
        win = sl[max(0, m.start() - 25): m.end() + 25]
        if re.search(r'\bisti(?:yor|yo|yro|yr)\b', win, flags=re.IGNORECASE):
            return norm_want(f"{base} istiyor")
        return base

    def _extract_cancel_intent(text: str) -> str:
        """
        Metinde abonelik iptal/kapatma niyeti varsa kısa bir etiket döndürür.
        Uydurma yok: sadece anahtar kelimelere göre 'İptal/Kapatma' döner.
        """
        if not text:
            return ""
        s = text.lower()

        # "iptal ettiricek" gibi yazımlar için kök yakalama yeterli
        if re.search(r"\biptal\b", s):
            return "İptal/Kapatma"

        # kapat/kapatacak/kapattır gibi
        if re.search(r"\bkapat\w*\b", s):
            return "İptal/Kapatma"

        # fesih / sonlandır
        if re.search(r"\bfesih\b", s) or re.search(r"\bsonland\w*\b", s) or re.search(r"\bsonlandır\w*\b", s):
            return "İptal/Kapatma"

        return ""

    parts: list[str] = []

    # Lokal / değer / sinyal kalıpları
    # Not: "lokelde", "lokelinde" gibi yazımlarda "lokal" alt dizisi olmayabilir; "lokel" de yakala
    _lokal_ctx = ("lokal" in t) or ("lokel" in t)

    def _lokal_ariza_sinyali() -> bool:
        """Teknik arıza sinyali: tek başına 'yok' (ör. 'evde yok') yanlış tetiklemesin."""
        if any_("değer", "deger", "db", "sinyal", "gelm") or any_("kötü", "bozuk"):
            return True
        # 'yok' sadece teknik bağlamda (evde yok / abone yok değil)
        if re.search(
            r"\b(?:değer|deger|sinyal|hat|port|kayıt|kayit|lokal|lokel)\w*\s+yok\b",
            t,
            flags=re.IGNORECASE,
        ):
            return True
        if re.search(r"\byok\s+gel", t) or re.search(r"gelm\w*\s+yok", t):
            return True
        return False

    # Olumlu durum: hat lokale sağlama alındı (arıza değil)
    if _lokal_ctx and (
        any_("sağlama", "saglam")
        and (any_("alındı", "alindi", "aldı", "aldi") or "sağlama al" in t or "saglam al" in t)
    ):
        if any_("evde yok") and any_("bilgi verildi", "bilgi veril"):
            parts.append("Lokalde sağlama alındı; abone evde yok, bilgi verildi")
        else:
            parts.append("Lokalde sağlama alındı")
    elif _lokal_ctx and _lokal_ariza_sinyali():
        if any_("kontrol", "bak", "erişim", "eris"):
            parts.append("Lokal değerleri bozuk, lokal kontrol gerekiyor")
        else:
            parts.append("Lokal kaynaklı sorun, kontrol gerekiyor")

    if any_("ulaşılam", "cevap vermiyor", "telefonlara cevap"):
        parts.append("Müşteriye ulaşılamıyor")

    # Adres güncelleme (metinden aynen çek)
    if any_("adres yanlış", "adresi yanlış", "adres yanlis", "adresi yanlis", "adres tutmuyor", "adres hatalı", "adres hatali", "doğru adres", "dogru adres"):
        addr = extract_address_snippet(clean_text)
        parts.append(f"Adres yanlış - {addr}" if addr else "Adres yanlış (güncellenecek)")
    if any_("modem arız", "modem ariz"):
        parts.append("Modem arızası şüphesi")
    if any_("port değiş", "port degis"):
        parts.append("Port değişikliği yapıldı")

    # Kablo/ankastre: gelecek zaman ile geçmiş zamanı ayır
    if any_("kablo", "ankastre") and any_("çek", "ceki", "çekil", "çektir", "cektir"):
        future = any_(
            "çekecek", "cek ecek", "çekilecek", "cekilecek", "çekicek", "cekicek",
            "beklen", "haber edecek", "bildirecek", "bilgi verecek", "yapacak",
            "çektirmesi", "cektirmesi", "çektir", "cektir", "gerekiyor", "gerekli"
        )
        past = any_(
            "çekildi", "cekildi", "çekilmiş", "cekmis", "çekilmiştir",
            "yapıldı", "yapildi", "tamamlandı", "tamamlandi", "değiştirildi", "degistirildi"
        )
        if future and not past:
            day_phrase = _extract_day_phrase(clean_text)
            if day_phrase:
                parts.append(f"Kablo çekimi bekleniyor - {day_phrase}")
            else:
                parts.append("Kablo çekimi bekleniyor")
        elif past:
            parts.append("Kablo/ankastre işlemi yapıldı")

    if any_("kablo", "ankastre") and any_("kop", "degis", "değiş"):
        parts.append("Kablo/ankastre işlemi yapıldı")

    # Mail: "atıldı/gönderildi" (geçmiş) ile "atalım/atılmalı" (gereken) ayrımı
    if "mail" in t or any_("mail at", "mail atar", "mail atabil"):
        sent = any_(
            "mail atıldı", "mail atildi", "mail gönderildi", "mail gonderildi",
            "mail yollandı", "mail yollandi", "mail iletildi", "mail atılmış", "mail atilmis"
        )
        # "mail atılmış/atıldı" metninde "mail at" geçtiği için false-positive olmasın:
        needed_explicit = any_(
            "mail atar", "mail atabil", "mail atalım", "mail atalim",
            "mail atılmalı", "mail atilmali", "mail atar mısınız", "mail atar misiniz",
            "mail atabilir misiniz"
        )
        needed_plain = re.search(r'\bmail\s+at\b', t) is not None
        needed = needed_explicit or (needed_plain and not sent)

        if sent and not needed_explicit:
            target = extract_mail_target(clean_text)
            reason = extract_mail_reason(clean_text)
            extra = " - ".join([x for x in [reason, target] if x])
            parts.append(f"Mail gönderildi - {extra}" if extra else "Mail gönderildi")
        elif needed:
            target = extract_mail_target(clean_text)
            why = extract_mail_needed_reason(clean_text)
            cancel = _extract_cancel_intent(clean_text)
            extra = " - ".join([x for x in [target, why, cancel] if x])
            parts.append(f"Mail ile bilgilendirme gerekiyor - {extra}" if extra else "Mail ile bilgilendirme gerekiyor")

    if parts:
        uniq: list[str] = []
        for p in parts:
            if p not in uniq:
                uniq.append(p)
        return " | ".join(uniq)

    if not clean_text:
        return ""
    # Kural eşleşmezse ham metin: Excel özet sütununda yarım kalmaması için limit biraz yüksek
    _fb = 220
    return (clean_text[:_fb].strip() + "…") if len(clean_text) > _fb else clean_text

def extract_address_snippet(text: str) -> str:
    """
    'Doğru adres ...' gibi kısımdan adresi mümkün olduğunca aynen çeker.
    Uydurma yok: metinde ne varsa onu alır.
    """
    if not text:
        return ""
    # Öncelik: 'doğru adres' sonrası
    m = re.search(r'(doğru adres|dogru adres)\s*[:\-]?\s*(.+)$', text, flags=re.IGNORECASE)
    if m:
        tail = m.group(2).strip()
        # çok uzamasın: ilk 120 karakter veya satır sonuna kadar
        tail = tail.split("\n", 1)[0].strip()
        if len(tail) > 140:
            tail = tail[:140].rstrip() + "…"
        return tail

    # Alternatif: 'adres' geçen kısım (örn: 'müşteri adresi yanlış. mustafa ... olacak')
    m2 = re.search(r'adres\w*\s*(?:yanlış|yanlis|hatalı|hatali)[^a-z0-9çğıöşü]*\s*(.+)$', text, flags=re.IGNORECASE)
    if m2:
        tail = m2.group(1).strip()
        tail = tail.split("\n", 1)[0].strip()
        if len(tail) > 140:
            tail = tail[:140].rstrip() + "…"
        return tail

    return ""

def extract_provider(text: str) -> str:
    """
    Mesaj içindeki ISS/marka anahtar kelimelerinden mailin nereye gideceğini tahmin eder.
    """
    if not text:
        return ""
    t = text.lower()
    providers = []

    def add(name: str):
        if name not in providers:
            providers.append(name)

    if "turknet" in t or "türkn" in t:
        add("Turknet")
    if "superonline" in t or "superonlıne" in t or "super on" in t:
        add("Superonline")
    if "digiturk" in t:
        add("Digiturk")
    if "dsmart" in t or "d-smart" in t or "d smart" in t:
        add("D-Smart")
    if "milenicom" in t:
        add("Milenicom")
    if "vodafone" in t:
        add("Vodafone")
    if "ttnet" in t or "türk telekom" in t or "turk telekom" in t or "telekom" in t:
        add("Türk Telekom")

    return ", ".join(providers)

def extract_mail_target(text: str) -> str:
    """
    'Mail at' geçen mesajlarda mailin nereye/kime atılacağını tahmin eder.
    Önce bilinen hedefleri (Turknet, NMS, ...) yakalar; yoksa 'X'e mail' kalıbını arar.
    """
    if not text:
        return ""

    t = text.lower()

    # 1) Açık hedefler
    explicit = []

    def add(name: str):
        if name and name not in explicit:
            explicit.append(name)

    # Bilinen ISS/markalar
    p = extract_provider(text)
    if p:
        for part in [x.strip() for x in p.split(",") if x.strip()]:
            add(part)

    # Operasyon hedefleri
    if re.search(r'\bnms\b', t) or "nms'e" in t or "nms e" in t or "nms’e" in t:
        add("NMS")
    if re.search(r'\bnoc\b', t) or "noc'a" in t or "noc a" in t or "noc’a" in t:
        add("NOC")
    if "saha ekibi" in t or "ekibe mail" in t:
        add("Saha ekibi")

    if explicit:
        return ", ".join(explicit)

    # 2) Genel kalıp:  "<hedef>'e mail"
    # örn: "nms'e mail atar mısınız", "firmaya mail atalım"
    m = re.search(r'\b([a-zçğıöşü0-9]{2,}(?:\s+[a-zçğıöşü0-9]{2,}){0,3})\s*(?:\'?e|\'?a|e|a|’e|’a)\s*mail\b', t, flags=re.IGNORECASE)
    if m:
        cand = m.group(1).strip()
        # çok genel kelimeleri normalize et
        if cand in ("firma", "firmaya", "iss", "ekip", "ekibe"):
            return cand.capitalize()
        return cand.upper() if cand in ("nms", "noc") else cand.capitalize()

    return ""

def extract_mail_reason(text: str) -> str:
    """
    'mail atılmış/gönderildi' cümlesinden, mailin nedeni gibi görünen kısa bir parçayı çıkarır.
    Örn: "atlak, konuyla ilgili mail atılmış." -> "atlak"
    """
    if not text:
        return ""

    # Cümle/parça bazlı ara
    chunks = re.split(r'[.!?\n]+', text)
    mail_chunk = None
    for c in chunks:
        ct = c.lower()
        if "mail" in ct and ("atıl" in ct or "gönder" in ct or "ilet" in ct or "yolla" in ct):
            mail_chunk = c.strip()
            break
    if not mail_chunk:
        return ""

    # mail kısmını temizle
    s = mail_chunk
    s = re.sub(r'\bmail\b', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\b(atılmış|atilmis|atıldı|atildi|gönderildi|gonderildi|iletildi|yollandı|yollandi)\b', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\b(konuyla ilgili|konu ile ilgili)\b', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip(" -,:;")

    # Çok uzamasın
    if len(s) > 60:
        s = s[:60].rstrip() + "…"

    return s

def extract_mail_needed_reason(text: str) -> str:
    """
    'mail atar mısınız / mail atalım' gibi durumlarda sebebi kısa çıkar.
    Örnek:
      - "kopma modem kaynaklı" -> "sorun modem kaynaklı"
      - "fiber dönüşüm kesilmesi gerekiyor" -> "fiber dönüşüm kesilmeli"
      - "hızını ... düşürmesi gerekiyor" -> "hız düşürülmeli"
    """
    if not text:
        return ""

    t = text.lower()

    if ("modem" in t) and ("kop" in t or "kopma" in t or "kaynaklı" in t or "arızalı" in t or "arizali" in t):
        return "sorun modem kaynaklı"

    if ("fiber" in t or "fibere" in t) and ("dönüşüm" in t or "donusum" in t) and ("kesil" in t):
        return "fiber dönüşüm kesilmeli"

    if ("hız" in t or "hiz" in t) and ("düş" in t or "dus" in t):
        return "hız düşürülmeli"

    return ""

def parse_ddmmyyyy(value: str) -> datetime:
    m = re.match(r'^\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*$', value or "")
    if not m:
        raise ValueError(f"Tarih formatı DD.MM.YYYY olmalı. Aldım: {value!r}")
    dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return datetime(yyyy, mm, dd, 0, 0, 0)

def build_range(from_str: str, to_str: str) -> tuple[datetime, datetime]:
    start = parse_ddmmyyyy(from_str)
    end_day = parse_ddmmyyyy(to_str)
    end = end_day.replace(hour=23, minute=59, second=59, microsecond=999000)
    return start, end

def extract_service_numbers(text):
    """
    Varsayılan: 10 haneli hizmet no (TR GSM 5xxxxxxxxx hariç).
    Ek: Mesajda 'pstn' geçiyorsa kısa PSTN no (genelde 6-8 hane) da yakala.
    """
    if not text:
        return []

    nums = re.findall(r'\b(?!5\d{9}\b)\d{10}\b', text)
    if "pstn" in text.lower():
        nums += re.findall(r'\b\d{6,8}\b', text)

    out = []
    for n in nums:
        if n not in out:
            out.append(n)
    return out

def strip_numbers(text, numbers):
    out = text or ""
    for n in numbers:
        out = out.replace(n, "")
    out = re.sub(r'\s+', ' ', out).strip()
    return out

def extract_mentions(text):
    if not text:
        return []
    found = []
    # @⁨Name⁩ formatı
    for m in re.finditer(r'@⁨([^⁩]+)⁩', text):
        v = (m.group(1) or "").strip()
        if v:
            found.append(v)
    # @Name formatı
    for m in re.finditer(r'@([^\s@]+)', text):
        v = (m.group(1) or "").strip()
        if not v:
            continue
        if re.match(r'^\+?\d', v):
            continue
        found.append(v)
    # uniq
    uniq = []
    for x in found:
        x = re.sub(r'[^\w\.\-\sçğıöşüÇĞİÖŞÜ]', ' ', x).strip()
        if x and x not in uniq and len(x) <= 60:
            uniq.append(x)

    def norm(v: str) -> str:
        v = v.lower()
        v = re.sub(r'[_\-.]+', ' ', v)
        v = re.sub(r'\s+', ' ', v).strip()
        return v

    # Kısa/uzun tekrarları temizle: "Selin" + "Selin Abla TT" → sadece uzun kalsın
    sorted_names = sorted(uniq, key=lambda v: len(norm(v)), reverse=True)
    kept = []
    for name in sorted_names:
        n = norm(name)
        redundant = False
        for k in kept:
            kn = norm(k)
            if kn.find(n) != -1 or n.find(kn) != -1:
                redundant = True
                break
        if not redundant:
            kept.append(name)
    return kept

def parse_whatsapp_txt(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.read().replace('\r\n', '\n').split('\n')

    # Desteklenen formatlar:
    # A) [7.04.2026 11:37:49] İsim: Mesaj
    # B) [11:07, 13.04.2026] İsim: Mesaj  (bazı exportlar)
    p_a = re.compile(r'^\[(\d{1,2})\.(\d{1,2})\.(\d{4})\s+(\d{1,2}):(\d{2})(?::(\d{2}))?\]\s*(.*)$')
    p_b = re.compile(r'^\[(\d{2}):(\d{2}),\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\]\s*(.*)$')

    messages = []
    current = None  # dict: dt(datetime), sender(str), text_parts(list)

    def flush():
        nonlocal current
        if not current:
            return
        text = "\n".join(current["text_parts"]).strip()
        sender = current["sender"].strip()

        service_numbers = extract_service_numbers(text)
        if not service_numbers:
            current = None
            return

        dt = current["dt"]
        dt_str = dt.strftime("%d.%m.%Y %H:%M")
        clean_desc = build_clean_description(text, service_numbers)
        summary = build_rule_summary(clean_desc, mentions=None)

        msg_id = hashlib.md5(f"{dt.isoformat()}{sender}{text}".encode("utf-8")).hexdigest()[:8]

        for num in service_numbers:
            messages.append({
                'hizmet_no': num,
                'ozet': summary,
                'yazan_kisi': sender,
                'tarih': dt_str,
                'ham_metin': text,
                'msg_id': msg_id
            })
        current = None

    for line in lines:
        line = re.sub(r'^[\u200e\u200f\u202a\u202b\u202c\u2066\u2067\u2068\u2069\ufeff]+', '', line)
        m = p_a.match(line)
        if m:
            flush()
            dd, mm, yyyy, hh, mi, ss, rest = m.groups()
            ss = ss or "00"
            dt = datetime(int(yyyy), int(mm), int(dd), int(hh), int(mi), int(ss))
            # sender/text split
            sender = ""
            text0 = rest or ""
            if ": " in text0:
                sender, text0 = text0.split(": ", 1)
            current = {"dt": dt, "sender": sender, "text_parts": [text0]}
            continue

        m = p_b.match(line)
        if m:
            flush()
            hh, mi, dd, mm, yyyy, rest = m.groups()
            dt = datetime(int(yyyy), int(mm), int(dd), int(hh), int(mi), 0)
            sender = ""
            text0 = rest or ""
            if ": " in text0:
                sender, text0 = text0.split(": ", 1)
            current = {"dt": dt, "sender": sender, "text_parts": [text0]}
            continue

        # continuation line
        if current is not None:
            current["text_parts"].append(line)

    flush()
    return messages

def process_files(filepaths: list[str], from_str: str | None = None, to_str: str | None = None) -> list[dict]:
    start = end = None
    if from_str and to_str:
        start, end = build_range(from_str, to_str)

    all_rows: list[dict] = []
    seen: set[str] = set()

    for fp in filepaths:
        for row in parse_whatsapp_txt(fp):
            # row['tarih'] is string; rebuild dt for filtering using the embedded msg_id? We have no dt.
            # Instead, parse from row['tarih'] (DD.MM.YYYY HH:MM).
            try:
                dt = datetime.strptime(row["tarih"], "%d.%m.%Y %H:%M")
            except Exception:
                dt = None

            if start and end and dt:
                if not (dt >= start and dt <= end):
                    continue

            # Çoklu dosya birleştirmede dedup
            dedup_key = f'{row.get("msg_id","")}::{row.get("hizmet_no","")}'
            if dedup_key in seen:
                continue
            seen.add(dedup_key)
            all_rows.append(row)

    return all_rows

def write_excel(messages, output_path):
    wb = Workbook()

    # ── Sheet 1: Özet ──
    ws1 = wb.active
    ws1.title = "Özet"

    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Hizmet No", "Özet", "Yazan Kişi", "Tarih", "Durum"]
    col_widths = [15, 70, 22, 18, 15]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.row_dimensions[1].height = 20

    for i, msg in enumerate(messages, 2):
        row_fill = alt_fill if i % 2 == 0 else None
        values = [msg['hizmet_no'], msg['ozet'], msg['yazan_kisi'], msg['tarih'], '']
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if row_fill:
                cell.fill = row_fill

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:E{len(messages)+1}"

    try:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        # Windows konsolda bazı kod sayfalarında unicode basmak hata verebiliyor
        print(f"{len(messages)} kayit yazildi -> {output_path}")
    except PermissionError:
        # Excel dosyası açıksa Windows'ta yazma engellenir
        base, ext = (output_path.rsplit(".", 1) + ["xlsx"])[:2]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{ts}.{ext}"
        wb.save(alt)
        print(f"{len(messages)} kayit yazildi -> {alt} (dosya kilitliydi, yeni isimle)")

# ── Test verisi ──
TEST_DATA = """[11:07, 13.04.2026] Ümit Abi TT 613: 8800328984 nolu abonenin arızası giderilemedi 8 SD A 1 .8 lokeline degeler kötü gelmekte bazı yerlerine değerler gelmemektedir lokelin kontrol edilmesi gereklidir @TT Nadir Abi @+90 543 575 21 50
[11:33, 13.04.2026] Ümit Abi TT 613: 8849634987 nolu arıza müşteri Muğla da hastanede olduğunu öğleden sonra 2 gibi burda olacağını belirtti @TT Nadir Abi
[12:01, 13.04.2026] Yiğit Abi TT: 8824439510 nolı arıza modem arızalı tivivu arıza kaydı açılıcak öğleden sonra değişim yapılacak @Anıl TT iys hata veriyor cihaz sorgulaması port değişimi yapılamıyor @Mehmet Abi TT @TT Nadir Abi
[12:16, 13.04.2026] Yiğit Abi TT: 8865240412 nolu diğer ıss arıza milenicom port değiştirildi abone 2 kere erkene geldi ping sorunu ve hat bağlantısını yapmamıştı firma mail atalım sürekli tekrar ediyor @Mehmet Abi TT @Anıl TT @TT Nadir Abi
"""

if __name__ == "__main__":
    # Kullanım:
    #   python test.py --from 13.04.2026 --to 14.04.2026 --out output\\ariza_takip.xlsx file1.txt file2.txt
    #   python test.py file1.txt
    args = sys.argv[1:]
    from_arg = to_arg = None
    out_arg = r"output\ariza_takip.xlsx"
    files = []

    i = 0
    while i < len(args):
        a = args[i]
        if a == "--from":
            from_arg = args[i + 1]; i += 2; continue
        if a == "--to":
            to_arg = args[i + 1]; i += 2; continue
        if a in ("--out", "-o"):
            out_arg = args[i + 1]; i += 2; continue
        if a in ("--help", "-h"):
            print("Kullanim:\n  python test.py --from DD.MM.YYYY --to DD.MM.YYYY --out output\\ariza_takip.xlsx file1.txt file2.txt\n  python test.py file1.txt")
            raise SystemExit(0)
        files.append(a)
        i += 1

    if not files:
        # Test modu
        import tempfile
        tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8')
        tmp.write(TEST_DATA)
        tmp.close()
        files = [tmp.name]
        try:
            messages = process_files(files, from_arg, to_arg)
        finally:
            os.unlink(tmp.name)
    else:
        messages = process_files(files, from_arg, to_arg)

    write_excel(messages, out_arg)